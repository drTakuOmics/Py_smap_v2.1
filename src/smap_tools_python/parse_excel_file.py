"""Utilities for extracting records from Excel spreadsheets."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Sequence
import math

try:  # pragma: no cover - optional dependency
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - openpyxl might not be installed
    load_workbook = None


def _header_from_rows(row1: Sequence[Any], row2: Sequence[Any]) -> List[str]:
    headers: List[str] = []
    for a, b in zip(row1, row2):
        a = "" if a is None else str(a)
        b = "" if b is None else str(b)
        if b and b.lower() != "nan":
            headers.append(f"{a}.{b}")
        else:
            headers.append(a)
    return headers


def parse_excel_file(path: str | Path, id_value: Any | None = None) -> List[Dict[str, Any]] | Dict[str, Any]:
    """Parse an Excel file into dictionaries.

    Parameters
    ----------
    path : str or Path
        Excel ``.xlsx`` file to read.
    id_value : optional
        If provided, only the row whose first column matches ``id_value`` is
        returned as a single dictionary.

    Returns
    -------
    list of dict or dict
        Parsed records. When ``id_value`` is given, a single dictionary is
        returned (or an empty dict if the ``id_value`` is not found).
    """

    if load_workbook is None:
        raise ImportError("openpyxl is required for parse_excel_file")

    wb = load_workbook(filename=Path(path), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {} if id_value is not None else []

    headers = _header_from_rows(rows[0], rows[1])
    records: List[Dict[str, Any]] = []
    for row in rows[2:]:
        rec: Dict[str, Any] = {}
        for key, val in zip(headers, row):
            if isinstance(val, float) and math.isnan(val):
                rec[key] = None
            elif isinstance(val, str) and val == "NaN":
                rec[key] = None
            else:
                rec[key] = val
        records.append(rec)

    if id_value is not None:
        key0 = headers[0] if headers else None
        for rec in records:
            if str(rec.get(key0)) == str(id_value):
                return rec
        return {}

    return records


# MATLAB-style alias
parseExcelFile = parse_excel_file


__all__ = ["parse_excel_file", "parseExcelFile"]

